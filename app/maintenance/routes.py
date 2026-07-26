import os
import uuid
from datetime import date, datetime, timedelta

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import current_user, login_required
from werkzeug.utils import secure_filename

from ..extensions import db
from ..models import ControlType, Machine, MaintenanceEvent, MaintenancePlan, Technician

maintenance_bp = Blueprint("maintenance", __name__, url_prefix="/maintenance")

ALLOWED_PHOTO_EXT = {".png", ".jpg", ".jpeg"}
MAX_PHOTO_SIZE = 10 * 1024 * 1024  # 10 MB
UPLOAD_SUBDIR = os.path.join("uploads", "manutenzioni")


def owner_only():
    if not current_user.is_owner:
        abort(403)


def _upload_folder():
    path = os.path.join(current_app.static_folder, UPLOAD_SUBDIR)
    os.makedirs(path, exist_ok=True)
    return path


def _save_photo(file_storage):
    """Salva una foto caricata e restituisce il percorso relativo per static/, o None."""
    if not file_storage or not file_storage.filename:
        return None
    ext = os.path.splitext(secure_filename(file_storage.filename))[1].lower()
    if ext not in ALLOWED_PHOTO_EXT:
        flash(f"Formato foto .{ext.lstrip('.')} non supportato: usa PNG o JPG.", "error")
        return None
    payload = file_storage.read()
    if not payload or len(payload) > MAX_PHOTO_SIZE:
        flash("La foto è vuota o supera il limite di 10 MB.", "error")
        return None
    filename = f"{uuid.uuid4().hex}{ext}"
    with open(os.path.join(_upload_folder(), filename), "wb") as out:
        out.write(payload)
    return f"{UPLOAD_SUBDIR}/{filename}".replace("\\", "/")


def _spawn_next_plan(plan):
    """Alla chiusura di un intervento ricorrente, crea il piano per il prossimo ciclo.

    Il piano appena concluso resta con stato "concluso" come traccia storica
    (utile per stampare la scheda e per l'audit ISO 9001); il nuovo piano
    riparte da "da programmare" con la scadenza ricalcolata.
    """
    nuovo = MaintenancePlan(
        machine_id=plan.machine_id,
        control_type_id=plan.control_type_id,
        technician_id=plan.technician_id,
        location=plan.location,
        contractor_name=plan.contractor_name,
        frequency_days=plan.frequency_days,
        next_due=date.today() + timedelta(days=plan.frequency_days),
        photo_instructions=plan.photo_instructions,
        status="da_programmare",
    )
    db.session.add(nuovo)
    return nuovo


@maintenance_bp.route("/")
@login_required
def dashboard():
    owner_only()
    today = date.today()
    critical = (
        MaintenancePlan.query.filter(
            MaintenancePlan.status != "concluso",
            MaintenancePlan.next_due <= today + timedelta(days=7),
        )
        .order_by(MaintenancePlan.next_due)
        .all()
    )
    machines = Machine.query.order_by(Machine.name).all()
    return render_template("maintenance/dashboard.html", critical=critical, machines=machines, today=today)


@maintenance_bp.route("/pipeline")
@login_required
def pipeline():
    owner_only()
    plans = MaintenancePlan.query.order_by(MaintenancePlan.next_due).all()
    columns = {
        "da_programmare": [p for p in plans if p.status == "da_programmare"],
        "in_esecuzione": [p for p in plans if p.status == "in_esecuzione"],
        "concluso": [p for p in plans if p.status == "concluso"],
    }
    return render_template("maintenance/pipeline.html", columns=columns)


@maintenance_bp.post("/pipeline/<int:plan_id>/status")
@login_required
def update_status(plan_id):
    owner_only()
    plan = db.get_or_404(MaintenancePlan, plan_id)
    new_status = request.form.get("status", "")
    if new_status not in {"da_programmare", "in_esecuzione", "concluso"}:
        flash("Stato non valido.", "error")
        return redirect(url_for("maintenance.pipeline"))

    if new_status == "in_esecuzione":
        location = request.form.get("location")
        contractor = request.form.get("contractor_name", "").strip()
        if location in {"interna", "fornitore"}:
            plan.location = location
        plan.contractor_name = contractor or None
        plan.status = "in_esecuzione"
    elif new_status == "concluso":
        note = request.form.get("note", "").strip()
        photo_path = _save_photo(request.files.get("photo"))
        db.session.add(
            MaintenanceEvent(
                machine_id=plan.machine_id,
                plan_id=plan.id,
                technician_id=plan.technician_id,
                control_type_name=plan.control_type.name,
                outcome=request.form.get("outcome", "eseguito"),
                note=note,
                photo_path=photo_path,
                performed_at=datetime.utcnow(),
            )
        )
        plan.status = "concluso"
        nuovo = _spawn_next_plan(plan)
        flash(f"Intervento registrato per {plan.machine.name}. Prossima scadenza: {nuovo.next_due.strftime('%d/%m/%Y')}.", "success")
    else:
        plan.status = "da_programmare"

    db.session.commit()
    return redirect(url_for("maintenance.pipeline"))


@maintenance_bp.route("/machines/<int:machine_id>")
@login_required
def machine_detail(machine_id):
    owner_only()
    machine = db.get_or_404(Machine, machine_id)
    technicians = Technician.query.order_by(Technician.name).all()
    control_types = ControlType.query.order_by(ControlType.name).all()
    return render_template(
        "maintenance/machine_detail.html",
        machine=machine,
        history=machine.events,
        technicians=technicians,
        control_types=control_types,
        today=date.today(),
    )


@maintenance_bp.post("/machines/<int:machine_id>/events")
@login_required
def add_event(machine_id):
    owner_only()
    machine = db.get_or_404(Machine, machine_id)
    control_type_name = request.form.get("control_type_name", "").strip() or "Intervento manuale"
    technician_id = request.form.get("technician_id") or None
    note = request.form.get("note", "").strip()
    photo_path = _save_photo(request.files.get("photo"))
    db.session.add(
        MaintenanceEvent(
            machine_id=machine.id,
            technician_id=int(technician_id) if technician_id else None,
            control_type_name=control_type_name,
            outcome=request.form.get("outcome", "eseguito"),
            note=note,
            photo_path=photo_path,
            performed_at=datetime.utcnow(),
        )
    )
    db.session.commit()
    flash("Intervento aggiunto al registro storico.", "success")
    return redirect(url_for("maintenance.machine_detail", machine_id=machine.id))


@maintenance_bp.route("/calendar")
@login_required
def calendar_view():
    owner_only()
    return render_template("maintenance/calendar.html")


@maintenance_bp.route("/api/due")
@login_required
def api_due():
    owner_only()
    plans = MaintenancePlan.query.filter(MaintenancePlan.status != "concluso").all()
    today = date.today()
    events = [
        {
            "id": plan.machine_id,
            "title": f"{plan.machine.name} ({plan.control_type.name})",
            "start": plan.next_due.isoformat(),
            "color": "#a52b2b" if plan.next_due < today else "#236b4a",
        }
        for plan in plans
    ]
    return jsonify(events)


@maintenance_bp.route("/plans/<int:plan_id>/print")
@login_required
def print_sheet(plan_id):
    owner_only()
    plan = db.get_or_404(MaintenancePlan, plan_id)
    return render_template("maintenance/print_sheet.html", plan=plan)


@maintenance_bp.post("/import")
@login_required
def import_excel():
    owner_only()
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Seleziona prima un file Excel (.xlsx).", "error")
        return redirect(url_for("maintenance.dashboard"))

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Il file è vuoto.")
        headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

        def col(row, name, default=""):
            if name not in headers:
                return default
            value = row[headers.index(name)]
            return value if value is not None else default

        importati = 0
        for row in rows[1:]:
            if not any(row):
                continue
            code = str(col(row, "codice")).strip()
            if not code:
                continue

            machine = Machine.query.filter_by(code=code).first()
            if not machine:
                machine = Machine(code=code, name=str(col(row, "nome_macchina", code)).strip())
                db.session.add(machine)
                db.session.flush()

            tech_name = str(col(row, "incaricato")).strip()
            technician = None
            if tech_name:
                technician = Technician.query.filter_by(name=tech_name).first()
                if not technician:
                    technician = Technician(name=tech_name)
                    db.session.add(technician)
                    db.session.flush()

            control_name = str(col(row, "tipologia", "Controllo generico")).strip()
            control_type = ControlType.query.filter_by(name=control_name).first()
            if not control_type:
                control_type = ControlType(name=control_name)
                db.session.add(control_type)
                db.session.flush()

            due_raw = col(row, "scadenza")
            if isinstance(due_raw, datetime):
                due = due_raw.date()
            elif isinstance(due_raw, date):
                due = due_raw
            elif due_raw:
                due = datetime.strptime(str(due_raw).strip(), "%d/%m/%Y").date()
            else:
                due = date.today() + timedelta(days=int(col(row, "frequenza", 30) or 30))

            plan = MaintenancePlan(
                machine_id=machine.id,
                control_type_id=control_type.id,
                technician_id=technician.id if technician else None,
                frequency_days=int(col(row, "frequenza", 30) or 30),
                next_due=due,
                location=str(col(row, "locazione", "interna")).strip().lower() or "interna",
                contractor_name=str(col(row, "ditta_esterna", "")).strip() or None,
                status="da_programmare",
            )
            db.session.add(plan)
            importati += 1

        db.session.commit()
        flash(f"Importazione completata: {importati} piani caricati.", "success")
    except Exception as exc:  # noqa: BLE001 - vogliamo mostrare l'errore reale all'utente
        db.session.rollback()
        flash(f"Errore durante l'importazione: {exc}", "error")

    return redirect(url_for("maintenance.dashboard"))


@maintenance_bp.post("/machines")
@login_required
def create_machine():
    owner_only()
    code = request.form.get("code", "").strip()
    name = request.form.get("name", "").strip()
    if not code or not name:
        flash("Codice e nome macchina sono obbligatori.", "error")
        return redirect(url_for("maintenance.dashboard"))
    if Machine.query.filter_by(code=code).first():
        flash("Esiste già un macchinario con questo codice.", "error")
        return redirect(url_for("maintenance.dashboard"))
    db.session.add(Machine(code=code, name=name))
    db.session.commit()
    flash(f"Macchinario {name} aggiunto.", "success")
    return redirect(url_for("maintenance.dashboard"))


@maintenance_bp.post("/machines/<int:machine_id>/plans")
@login_required
def create_plan(machine_id):
    owner_only()
    machine = db.get_or_404(Machine, machine_id)
    control_name = request.form.get("control_type_name", "").strip()
    if not control_name:
        flash("Indica il tipo di controllo/intervento.", "error")
        return redirect(url_for("maintenance.machine_detail", machine_id=machine.id))

    control_type = ControlType.query.filter_by(name=control_name).first()
    if not control_type:
        control_type = ControlType(name=control_name)
        db.session.add(control_type)
        db.session.flush()

    frequency = int(request.form.get("frequency_days", 30) or 30)
    technician_id = request.form.get("technician_id") or None

    db.session.add(
        MaintenancePlan(
            machine_id=machine.id,
            control_type_id=control_type.id,
            technician_id=int(technician_id) if technician_id else None,
            frequency_days=frequency,
            next_due=date.today() + timedelta(days=frequency),
            status="da_programmare",
        )
    )
    db.session.commit()
    flash("Piano di manutenzione creato.", "success")
    return redirect(url_for("maintenance.machine_detail", machine_id=machine.id))
