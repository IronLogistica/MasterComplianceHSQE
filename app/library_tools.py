"""Conversione dei documenti della libreria (docx/xlsx) in HTML, per la
visualizzazione e la compilazione "paperless" dentro l'applicazione.

I file .doc/.xls legacy non ancora convertiti in .docx/.xlsx non sono
supportati da questa funzione: restano scaricabili ma non visualizzabili
in-app (document_to_html restituisce None in quel caso).
"""

import os

import mammoth
from openpyxl import load_workbook


def _docx_to_html(abs_path):
    with open(abs_path, "rb") as f:
        result = mammoth.convert_to_html(f)
    return result.value


def _xlsx_to_html(abs_path):
    wb = load_workbook(abs_path, data_only=True)
    ws = wb.active
    rows_html = []
    for row in ws.iter_rows(values_only=True):
        cells = "".join(f"<td>{'' if c is None else c}</td>" for c in row)
        rows_html.append(f"<tr>{cells}</tr>")
    return "<table class='doc-table'>" + "".join(rows_html) + "</table>"


def document_to_html(abs_path):
    """Converte un documento in HTML, o None se il formato non e' supportato."""
    ext = os.path.splitext(abs_path)[1].lower()
    try:
        if ext == ".docx":
            return _docx_to_html(abs_path)
        if ext == ".xlsx":
            return _xlsx_to_html(abs_path)
    except Exception:
        return None
    return None
